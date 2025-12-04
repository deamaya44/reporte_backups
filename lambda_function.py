import json
import boto3
import os
from datetime import datetime, timedelta, timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Configuración
BUCKET_NAME = os.environ['BUCKET_NAME']
FROM_EMAIL = os.environ['FROM_EMAIL']
TO_EMAILS = os.environ['TO_EMAILS'].split(',')
CC_EMAILS = os.environ.get('CC_EMAILS', '').split(',') if os.environ.get('CC_EMAILS') else []

def get_account_name(account_id, org_client):
    """Obtiene el nombre de la cuenta desde AWS Organizations"""
    try:
        response = org_client.describe_account(AccountId=account_id)
        return response['Account']['Name']
    except Exception as e:
        print(f"No se pudo obtener nombre para cuenta {account_id}: {str(e)}")
        return account_id

def lambda_handler(event, context):
    """Función principal de Lambda"""
    
    backup_client = boto3.client('backup')
    s3_client = boto3.client('s3')
    ses_client = boto3.client('ses')
    org_client = boto3.client('organizations')
    
    # Cache para nombres de cuentas
    account_names_cache = {}
    
    # Fechas simples - solo día completo
    now = datetime.now(timezone.utc)
    today = now.date()
    yesterday = today - timedelta(days=1)
    
    print(f"Fecha de hoy: {today}")
    print(f"Fecha de ayer: {yesterday}")
    
    # Obtener todos los backup jobs (últimos 2 días)
    all_jobs = []
    paginator = backup_client.get_paginator('list_backup_jobs')
    
    # Filtrar por los últimos 2 días para traer menos datos
    by_created_after = datetime.combine(yesterday, datetime.min.time(), tzinfo=timezone.utc)
    
    for page in paginator.paginate(
        ByAccountId='*',
        ByCreatedAfter=by_created_after
    ):
        all_jobs.extend(page['BackupJobs'])
    
    print(f"Total de jobs obtenidos: {len(all_jobs)}")
    
    # Filtrar por fecha simple - solo comparar el día
    yesterday_jobs = []
    today_jobs = []
    
    for job in all_jobs:
        creation_date = job.get('CreationDate')
        
        if not creation_date:
            continue
        
        # Obtener solo la fecha (sin hora)
        job_date = creation_date.date()
        
        if job_date == yesterday:
            yesterday_jobs.append(job)
        elif job_date == today:
            today_jobs.append(job)
    
    print(f"Jobs de ayer ({yesterday}): {len(yesterday_jobs)}")
    print(f"Jobs de hoy ({today}): {len(today_jobs)}")
    
    # Generar resúmenes con nombres de cuentas dinámicos
    yesterday_summary = generate_summary(yesterday_jobs, org_client, account_names_cache)
    today_summary = generate_summary(today_jobs, org_client, account_names_cache)
    
    # Crear archivo Excel
    excel_buffer = create_excel_report(yesterday_jobs, today_jobs, yesterday, today, org_client, account_names_cache)
    
    # Guardar en S3
    filename = f"backup-report-{today.strftime('%Y-%m-%d')}.xlsx"
    s3_key = f"reports/{today.strftime('%Y/%m')}/{filename}"
    
    s3_client.put_object(
        Bucket=BUCKET_NAME,
        Key=s3_key,
        Body=excel_buffer.getvalue(),
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    print(f"Reporte guardado en S3: s3://{BUCKET_NAME}/{s3_key}")
    
    # Enviar email
    send_email(
        ses_client,
        yesterday_summary,
        today_summary,
        excel_buffer,
        filename,
        yesterday,
        today
    )
    
    return {
        'statusCode': 200,
        'body': json.dumps({
            'message': 'Reporte generado y enviado exitosamente',
            's3_location': f"s3://{BUCKET_NAME}/{s3_key}",
            'yesterday_jobs': len(yesterday_jobs),
            'today_jobs': len(today_jobs)
        })
    }

def generate_summary(jobs, org_client, account_names_cache):
    """Genera resumen por cuenta y estado"""
    summary = {}
    
    for job in jobs:
        account_id = job.get('AccountId', 'Unknown')
        
        # Obtener nombre de cuenta (con cache)
        if account_id not in account_names_cache:
            account_names_cache[account_id] = get_account_name(account_id, org_client)
        
        account_name = account_names_cache[account_id]
        state = job.get('State', 'UNKNOWN')
        
        if account_name not in summary:
            summary[account_name] = {'COMPLETED': 0, 'FAILED': 0, 'OTHER': 0}
        
        if state == 'COMPLETED':
            summary[account_name]['COMPLETED'] += 1
        elif state == 'FAILED':
            summary[account_name]['FAILED'] += 1
        else:
            summary[account_name]['OTHER'] += 1
    
    return summary

def create_excel_report(yesterday_jobs, today_jobs, yesterday_date, today_date, org_client, account_names_cache):
    """Crea archivo Excel con el reporte detallado"""
    wb = Workbook()
    
    # Configurar estilos
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Hoja 1: Resumen
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    
    # Headers
    headers = ["AccountName", "T. COMPLETED", "T. FAILED"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Datos del resumen (día actual)
    today_summary = generate_summary(today_jobs, org_client, account_names_cache)
    row = 2
    for account_name in sorted(today_summary.keys()):
        ws_summary.cell(row=row, column=1, value=account_name)
        ws_summary.cell(row=row, column=2, value=today_summary[account_name]['COMPLETED'])
        ws_summary.cell(row=row, column=3, value=today_summary[account_name]['FAILED'])
        row += 1
    
    # Ajustar anchos
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    
    # Hoja 2: Detalle Día Anterior
    ws_yesterday = wb.create_sheet("Día Anterior")
    create_detail_sheet(ws_yesterday, yesterday_jobs, header_fill, header_font, org_client, account_names_cache)
    
    # Hoja 3: Detalle Día Actual
    ws_today = wb.create_sheet("Día Actual")
    create_detail_sheet(ws_today, today_jobs, header_fill, header_font, org_client, account_names_cache)
    
    # Guardar en buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def create_detail_sheet(worksheet, jobs, header_fill, header_font, org_client, account_names_cache):
    """Crea hoja de detalle con todos los jobs"""
    headers = [
        "BackupJobID", "Status", "AccountID", "AccountName", 
        "ResourceName", "MessageCategory", "ResourceID", 
        "ResourceType", "CreationTime"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    row = 2
    for job in jobs:
        account_id = job.get('AccountId', '-')
        
        # Obtener nombre de cuenta (con cache)
        if account_id not in account_names_cache:
            account_names_cache[account_id] = get_account_name(account_id, org_client)
        
        account_name = account_names_cache[account_id]
        
        worksheet.cell(row=row, column=1, value=job.get('BackupJobId', '-'))
        worksheet.cell(row=row, column=2, value=job.get('State', '-'))
        worksheet.cell(row=row, column=3, value=account_id)
        worksheet.cell(row=row, column=4, value=account_name)
        worksheet.cell(row=row, column=5, value=job.get('ResourceName', '-'))
        worksheet.cell(row=row, column=6, value=job.get('StatusMessage', 'Success'))
        worksheet.cell(row=row, column=7, value=job.get('ResourceArn', '-'))
        worksheet.cell(row=row, column=8, value=job.get('ResourceType', '-'))
        
        creation_date = job.get('CreationDate')
        if creation_date:
            worksheet.cell(row=row, column=9, value=creation_date.strftime('%Y-%m-%d %H:%M:%S'))
        else:
            worksheet.cell(row=row, column=9, value='-')
        
        row += 1
    
    # Ajustar anchos
    for col in range(1, 10):
        worksheet.column_dimensions[chr(64 + col)].width = 15

def send_email(ses_client, yesterday_summary, today_summary, excel_buffer, filename, yesterday_date, today_date):
    """Envía el email con el reporte"""
    
    # Crear mensaje
    msg = MIMEMultipart()
    msg['From'] = FROM_EMAIL
    msg['To'] = ', '.join(TO_EMAILS)
    if CC_EMAILS and CC_EMAILS[0]:
        msg['Cc'] = ', '.join(CC_EMAILS)
    msg['Subject'] = f"HV - Reporte de Backups de Máquinas Virtuales AWS – Resumen {today_date}"
    
    # Cuerpo del email
    body = generate_email_body(yesterday_summary, today_summary, yesterday_date, today_date)
    msg.attach(MIMEText(body, 'html'))
    
    # Adjuntar Excel
    excel_buffer.seek(0)
    attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    attachment.set_payload(excel_buffer.read())
    encoders.encode_base64(attachment)
    attachment.add_header('Content-Disposition', f'attachment; filename={filename}')
    msg.attach(attachment)
    
    # Preparar destinatarios
    destinations = TO_EMAILS.copy()
    if CC_EMAILS and CC_EMAILS[0]:
        destinations.extend(CC_EMAILS)
    
    # Enviar
    try:
        response = ses_client.send_raw_email(
            Source=FROM_EMAIL,
            Destinations=destinations,
            RawMessage={'Data': msg.as_bytes()}
        )
        print(f"Email enviado. MessageId: {response['MessageId']}")
    except Exception as e:
        print(f"Error enviando email: {str(e)}")
        raise

def generate_email_body(yesterday_summary, today_summary, yesterday_date, today_date):
    """Genera el cuerpo HTML del email"""
    
    # Resumen día anterior
    if yesterday_summary:
        yesterday_html = "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse;'>"
        yesterday_html += "<tr style='background-color: #0066CC; color: white;'>"
        yesterday_html += "<th>AccountName</th><th>T. COMPLETED</th><th>T. FAILED</th></tr>"
        for account_name in sorted(yesterday_summary.keys()):
            yesterday_html += f"<tr><td>{account_name}</td>"
            yesterday_html += f"<td>{yesterday_summary[account_name]['COMPLETED']}</td>"
            yesterday_html += f"<td>{yesterday_summary[account_name]['FAILED']}</td></tr>"
        yesterday_html += "</table>"
    else:
        yesterday_html = "No hay respaldos en este rango."
    
    # Resumen día actual
    if today_summary:
        today_html = "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse;'>"
        today_html += "<tr style='background-color: #0066CC; color: white;'>"
        today_html += "<th>AccountName</th><th>T. COMPLETED</th><th>T. FAILED</th></tr>"
        for account_name in sorted(today_summary.keys()):
            today_html += f"<tr><td>{account_name}</td>"
            today_html += f"<td>{today_summary[account_name]['COMPLETED']}</td>"
            today_html += f"<td>{today_summary[account_name]['FAILED']}</td></tr>"
        today_html += "</table>"
    else:
        today_html = "No hay respaldos en este rango."
    
    body = f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <p><strong>Resumen del DÍA ANTERIOR ({yesterday_date}):</strong></p>
        {yesterday_html}
        
        <br/>
        
        <p><strong>Resumen del DÍA ACTUAL ({today_date}):</strong></p>
        {today_html}
        
        <br/>
        <p>Se adjunta el reporte detallado en Excel.</p>
        
        <br/>
        <p>Saludos,<br/>
        Av. Apoquindo 5950, piso 4, edificio Wework Suite<br/>
        Santiago – Chile</p>
    </body>
    </html>
    """
    
    return body